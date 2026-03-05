import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import os

def create_formatted_template(filename="VehicularRecordsTemplate.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Define colors
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=12)
    small_white_font = Font(color="FFFFFF", size=9)
    
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True, size=10)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Setup Title Headers
    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")
    ws.merge_cells("A3:M3")
    ws.merge_cells("A4:M4")
    
    for row in range(1, 5):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = black_fill
    
    ws["A1"] = "Republic of the Philippines"
    ws["A1"].font = small_white_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A2"] = "Local Government Unit of Manolo Fortich"
    ws["A2"].font = small_white_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "GENERAL SERVICE OFFICE"
    ws["A3"].font = small_white_font
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "VEHICULAR RECORDS"
    ws["A4"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    # Set row height for the title area
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 30

    # Define columns
    columns = [
        "OFFICE",
        "PLATE NUMBER",
        "ENGINE NUMBER",
        "CHASSIS NO.",
        "BRAND/ BODY TYPE",
        "YEAR MODEL",
        "EXPIRATION DATE",
        "ACQUISITION COST",
        "ACQUISITION DATE",
        "ACCOUNTABLE PERSON",
        "STATUS"
    ]
    
    # Write columns
    for col_index, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=5, column=col_index)
        cell.fill = gray_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.value = column_name

    # Set column widths
    column_widths = {
        "A": 15,  # OFFICE
        "B": 20,  # PLATE NUMBER
        "C": 25,  # ENGINE NUMBER
        "D": 25,  # CHASSIS NO.
        "E": 25,  # BRAND/ BODY TYPE
        "F": 15,  # YEAR MODEL
        "G": 20,  # EXPIRATION DATE
        "H": 20,  # ACQUISITION COST
        "I": 20,  # ACQUISITION DATE
        "J": 30,  # ACCOUNTABLE PERSON
        "K": 15,  # STATUS
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add a row of dummy data
    dummy_data = [
        "MEO", "100101/7005", "E3W8E-057005", "PA0DE1110L0037399", "YAMAHA/MC", "2019", "2024-05-15", "", "", "RICHARD SITON", "SERVICEABLE"
    ]
    for col_index, value in enumerate(dummy_data, start=1):
        cell = ws.cell(row=6, column=col_index)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filename)
    print(f"Created template: {filename}")

if __name__ == "__main__":
    create_formatted_template()
